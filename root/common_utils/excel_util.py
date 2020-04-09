from openpyxl import Workbook
from root.common_utils.log_util import Logger
import traceback
from root.common_utils.file_util import create_file_if_not_exist


# wb = Workbook()

# https://www.cnblogs.com/programmer-tlh/p/10461353.html

# sheet = wb.create_sheet("Mysheet", 0)

# mydb = connector.connect(
#     host="192.168.64.128",
#     user="root",
#     passwd="65536",
#     database="mysql"
# )

# mycursor.execute("SHOW TABLES")
# for x in mycursor:
#     print(x)

# header = ["id", "username", "buy_name", "buy_mobile"]
#
# datas = []
# datas.append(header)

# mycursor = mydb.cursor()
# mycursor.execute("select * from t_dg_buy_user")
# all_data = mycursor.fetchall()
# fileds = [field[0] for field in mycursor.description]
#
# sheet.append(header)
#
# # 从第一行开始写
# for data in rows:
#     sheet.append(data)
#
# wb.save("balances.xlsx")

class ExcelGenerator(object):
    logger = Logger.get_instance()

    @staticmethod
    def generate_excel_file(header, rows, file_path):
        try:
            create_file_if_not_exist(file_path)
            wb = Workbook()
            sheet = wb.create_sheet("Mysheet", 0)
            sheet.append(header)

            for data in rows:
                if data:
                    sheet.append(data)

            wb.save(file_path)
            ExcelGenerator.logger.info("save excel " + file_path + " success")
            return True
        except Exception:
            ExcelGenerator.logger.error(traceback.format_exc())
            return False
